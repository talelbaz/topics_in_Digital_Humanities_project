# import geopandas
import csv

import geopy
import openpyxl


def get_coords(name):
    try:
        locator = geopy.Nominatim(user_agent="myGeocoder")
        location = locator.geocode(name)
        return [location.latitude, location.longitude]
    except:
        return []

    # print("Latitude = {}, Longitude = {}".format(location.latitude, location.longitude))


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    # import module

    # load excel with its path
    wrkbk = openpyxl.load_workbook("allMetaData.xlsx")

    sh = wrkbk.active

    # iterate through excel and display data
    with open('data_after_processing15.csv', 'w', newline='', encoding="utf-8") as csvfile:
        spamwriter = csv.writer(csvfile, delimiter=',',)
        # spamwriter.writerow(["latitude", "longitude", "a", "b", "c", "d"])
        for i in range(2, sh.max_row + 1):
            print(i)
            cell_obj = sh.cell(row=i, column=14)
            if cell_obj is not None:
                if cell_obj.value:
                    row = []
                    coords = get_coords(cell_obj.value)
                    row += coords
                    for j in range(1, sh.max_column + 1):
                        row.append(sh.cell(row=i, column=j).value)
                    # if all(row):
                    spamwriter.writerow(row)
